import csv
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .services.business_rules import authorization_narrative

CALL_HEADERS = [
    "Call date",
    "Authorization",
    "Facility",
    "Specialty",
    "Diagnosis",
    "Caller",
    "Result",
    "Recommendation",
]


def call_rows(queryset):
    for call in queryset.iterator(chunk_size=500):
        yield [
            call.call_at.isoformat(),
            call.authorization.authorization_number,
            call.facility.display_key,
            call.specialty.name,
            call.diagnosis.code,
            call.caller.username,
            call.result_phrase,
            call.recommendation,
        ]


def calls_csv_response(queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="provider-calls.csv"'
    writer = csv.writer(response)
    writer.writerow(CALL_HEADERS)
    writer.writerows(call_rows(queryset))
    return response


def calls_excel_response(queryset):
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Provider Calls"
    sheet.append(CALL_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F3D46")
        cell.alignment = Alignment(wrap_text=True)
    for row in call_rows(queryset):
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [22, 18, 38, 24, 14, 18, 58, 58]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="provider-calls.xlsx"'
    return response


def authorization_pdf_response(authorization):
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream, pagesize=LETTER, rightMargin=48, leftMargin=48, topMargin=48, bottomMargin=48
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Authorization Provider Summary", styles["Title"]),
        Spacer(1, 14),
        Paragraph(authorization_narrative(authorization), styles["BodyText"]),
        Spacer(1, 18),
    ]
    rows = [["Date", "Facility", "Outcome"]]
    for call in authorization.calls.select_related("facility").order_by("call_at"):
        rows.append([call.call_at.strftime("%Y-%m-%d"), call.facility.name, call.result_phrase])
    table = Table(rows, colWidths=[72, 150, 270], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3D46")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    story.append(table)
    document.build(story)
    response = HttpResponse(stream.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="authorization-{authorization.authorization_number}.pdf"'
    )
    return response


def report_excel_response(metrics):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Report Summary"
    summary.append(["Provider Availability Report", f"{metrics['period_start']} to {metrics['period_end']}"])
    summary.append([])
    summary.append(["Metric", "Value", "Definition"])
    definitions = [
        ("Calls", metrics["total_calls"], "All provider calls in the selected period"),
        (
            "Successful outcomes",
            metrics["successful_calls"],
            "Calls meeting guidelines, including urgent referral",
        ),
        (
            "Success rate",
            metrics["success_rate"] / 100,
            f"Successful calls / {metrics['success_denominator']} total calls",
        ),
        ("Unable to contact", metrics["unable_to_contact"], "Calls where no voicemail was left"),
        (
            "Unable-to-contact rate",
            metrics["unable_rate"] / 100,
            f"Unable-to-contact calls / {metrics['success_denominator']} total calls",
        ),
    ]
    for row in definitions:
        summary.append(row)
    summary[6][1].number_format = "0.0%"
    summary[8][1].number_format = "0.0%"
    for cell in summary[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F3D46")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 55
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="provider-availability-report.xlsx"'
    return response
